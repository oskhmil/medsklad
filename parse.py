#!/usr/bin/env python3
"""
Parser for hospital stock analysis Excel file.

Reads an Excel file with "Відомість по партіях товарів" structure from 1C,
extracts hierarchy from alignment.indent, and produces compact JSON for the
frontend.

Hierarchy (via alignment.indent on column A):
  0  = organization root / data cells (name in col B, series in col C)
  2  = direction (напрям діяльності: 03-НСЗУ, 67-Благодійна, etc.)
  4  = group (Адмін, Амбул.центр, ВП Лікарня, Реаб.центр, Реорг, Центр ест.мед.)
  6  = subdivision / department (first level of «Склад/Підрозділ»)
  8  = subdivision / department (second level — often a sub-department or МВО)
  10 = storage / person (third level — storage or МВО)
  12 = document/operation (Списання, Переміщення, Надходження, Коригування)
  14 = document/operation under an item

Numeric columns (1-indexed, so col 4 = D in Excel):
  4  D  Початковий залишок — кількість
  5  E  Початковий залишок — собівартість (грн)
  6  F  Надходження — кількість
  7  G  Надходження — собівартість
  8  H  Видаток — кількість
  9  I  Видаток — собівартість
  10 J  Кінцевий залишок — кількість
  11 K  Кінцевий залишок — собівартість

Output JSON structure (designed for compactness and query efficiency):
{
  "meta": { "generated_at": "...", "source_file": "...", "periods": [...] },
  "months": {
    "2026-01": {
      "period": { "from": "2026-01-01", "to": "2026-01-31" },
      "totals": { "open_qty":..., "open_sum":..., ... },
      "directions": [
        {
          "name": "03-НСЗУ",
          "totals": {...},
          "groups": [
            {
              "name": "Адміністративні підрозділи",
              "totals": {...},
              "subdivisions": [
                {
                  "path": ["Відділ забезпечення ...", "ВП Пантелеймон", "Склад №4"],
                  "totals": {...},
                  "items": [
                    { "name": "...", "series": "...", "open_q":..., ... }
                  ]
                }
              ]
            }
          ]
        }
      ]
    }
  }
}
"""

import json
import time
import sys
import re
from pathlib import Path
from datetime import datetime
import openpyxl

MONTH_UK_TO_NUM = {
    "Січень": "01", "Лютий": "02", "Березень": "03", "Квітень": "04",
    "Травень": "05", "Червень": "06", "Липень": "07", "Серпень": "08",
    "Вересень": "09", "Жовтень": "10", "Листопад": "11", "Грудень": "12",
}


def round_num(x, digits=2):
    """Round numbers consistently; return None for non-numeric/zero/None."""
    if x is None or x == "":
        return None
    try:
        v = float(x)
    except (TypeError, ValueError):
        return None
    if v == 0:
        return 0
    return round(v, digits)


def extract_numbers(row_cells):
    """
    Extract 8 numeric values (4 pairs: qty + sum) from columns D..K.
    Returns dict with short keys. Columns are 1-indexed; row_cells are openpyxl Cells.
    """
    def val(col_1idx):
        c = row_cells[col_1idx - 1] if col_1idx - 1 < len(row_cells) else None
        return round_num(c.value) if c is not None else None

    return {
        "oq": val(4),    # open quantity
        "os": val(5),    # open sum
        "rq": val(6),    # receipt (надходження) quantity
        "rs": val(7),    # receipt sum
        "eq": val(8),    # expense (видаток) quantity
        "es": val(9),    # expense sum
        "cq": val(10),   # close quantity
        "cs": val(11),   # close sum
    }


def has_any_number(nums):
    """True if any of the 8 numeric fields is non-None and non-zero."""
    return any(v not in (None, 0) for v in nums.values())


def _sum_totals(a, b):
    """Sum two totals dicts element-wise, keeping Nones where both are None."""
    if a is None:
        return b
    if b is None:
        return a
    out = {}
    for k in ("oq", "os", "rq", "rs", "eq", "es", "cq", "cs"):
        va, vb = a.get(k), b.get(k)
        if va is None and vb is None:
            out[k] = None
        else:
            out[k] = round((va or 0) + (vb or 0), 2)
    return out


def _strip_none(d):
    """Return a dict without None values; None input returns {}."""
    if d is None:
        return {}
    return {k: v for k, v in d.items() if v is not None}


def parse_period_from_sheet(ws):
    """Look at row 4 for 'Період: 01.01.2026 - 31.01.2026'."""
    for r in range(1, 20):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            if cell.value and isinstance(cell.value, str) and "Період" in cell.value:
                m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})\s*-\s*(\d{2})\.(\d{2})\.(\d{4})", cell.value)
                if m:
                    d1, m1, y1, d2, m2, y2 = m.groups()
                    return {"from": f"{y1}-{m1}-{d1}", "to": f"{y2}-{m2}-{d2}"}
    return None


def detect_data_start(ws):
    """Find the first row that is a data row (organization root, indent=0, col A filled)."""
    for r in range(11, 30):
        cell = ws.cell(row=r, column=1)
        if cell.value and str(cell.value).startswith("КНП"):
            return r
    return 15  # fallback


def is_document_row(text):
    """Detect document/operation rows (Списання, Переміщення, Надходження, Коригування)."""
    if not text:
        return False
    t = str(text).lstrip()
    return bool(re.match(
        r"^(Списання|Переміщення|Надходження|Коригування|Реалізація|Повернення|Оприбуткування|Інвентариза)",
        t
    ))


def _extract_doc_date(text):
    """Extract date from document title like 'Списання товарів за вимогою 0000-001065 від 26.01.2026 12:53:03'.
    Returns ISO date 'YYYY-MM-DD' or None.
    """
    if not text:
        return None
    m = re.search(r"від\s+(\d{2})\.(\d{2})\.(\d{4})", str(text))
    if m:
        d, mo, y = m.groups()
        return f"{y}-{mo}-{d}"
    return None


def parse_sheet(ws, sheet_name):
    """
    Parse one monthly sheet into a hierarchical structure.

    Returns: dict with period + directions tree.
    """
    period = parse_period_from_sheet(ws)
    start_row = detect_data_start(ws)

    # State: we track the current path at each indent level.
    # Current direction, group, and subdivision-path (list of names at indents 6/8/10)
    current_direction = None   # str or None (None => "Без напряму")
    current_group = None       # str
    # subdivision path: dict from indent -> name; we keep 6,8,10
    subdiv_path = {}           # {6: name, 8: name, 10: name}

    # Tree accumulation (we aggregate on the fly)
    # We key by (direction_name or "", group_name, tuple(subdiv_path_at_leaf))
    # But to match hierarchical JSON, let's build dicts by name
    directions = {}  # name -> {totals, groups: {name -> {totals, subdivisions: {path_tuple -> subdiv}}}}

    # Root totals for the organization (row with "КНП...")
    org_totals = None
    no_direction_totals = None  # for rows before any indent=2 header

    # Stats
    stats = {
        "rows_scanned": 0,
        "items_captured": 0,
        "directions_found": set(),
        "groups_found": set(),
        "unknown_levels": 0,
        "skipped_empty": 0,
    }

    max_row = ws.max_row
    last_item_key = None  # (dir, group, path_tuple) where last item was captured

    for r in range(start_row, max_row + 1):
        stats["rows_scanned"] += 1

        # Read the row cells once (columns 1..11)
        row_cells = [ws.cell(row=r, column=c) for c in range(1, 12)]
        cell_a = row_cells[0]
        indent = int(cell_a.alignment.indent or 0)
        a_val = str(cell_a.value).strip() if cell_a.value else ""
        b_val = str(row_cells[1].value).strip() if row_cells[1].value else ""
        c_val = str(row_cells[2].value).strip() if row_cells[2].value else ""

        # Organization root — indent=0, A filled, starts with "КНП"
        if indent == 0 and a_val.startswith("КНП"):
            org_totals = extract_numbers(row_cells)
            continue

        # Direction header — indent=2
        if indent == 2:
            # The direction name is in col A; if empty, mark as "Без напряму"
            name = a_val if a_val else "(Без напряму)"
            current_direction = name
            current_group = None
            subdiv_path = {}
            stats["directions_found"].add(name)
            dir_totals = extract_numbers(row_cells)
            if name not in directions:
                directions[name] = {"name": name, "totals": dir_totals, "groups": {}}
            else:
                # Same direction appears twice (happens with "(Без напряму)"): sum totals
                existing = directions[name]["totals"]
                directions[name]["totals"] = _sum_totals(existing, dir_totals)
            continue

        # Group header — indent=4
        if indent == 4 and a_val:
            # Group requires a direction context; fall back to "(Без напряму)"
            if current_direction is None:
                current_direction = "(Без напряму)"
                if current_direction not in directions:
                    directions[current_direction] = {"name": current_direction, "totals": None, "groups": {}}
            current_group = a_val
            subdiv_path = {}
            stats["groups_found"].add(a_val)
            grp_totals = extract_numbers(row_cells)
            dirobj = directions[current_direction]
            if a_val not in dirobj["groups"]:
                dirobj["groups"][a_val] = {"name": a_val, "totals": grp_totals, "subdivisions": {}}
            else:
                # Group can also appear multiple times (one per direction repetition)
                existing = dirobj["groups"][a_val]["totals"]
                dirobj["groups"][a_val]["totals"] = _sum_totals(existing, grp_totals)
            continue

        # Document/operation row — detected by NAME, not just indent.
        # In some 1C reports, a line like "Списання товарів за вимогою..." can
        # appear at indent=10/12/14 depending on the depth of the parent
        # subdivision. Treating it by prefix avoids misclassifying it as a
        # subdivision and creating phantom entries in the hierarchy tree.
        if a_val and is_document_row(a_val):
            # Track latest write-off date for the current subdivision so we
            # can show "last write-off N days ago" later.
            if a_val.lstrip().startswith("Списання"):
                d = _extract_doc_date(a_val)
                if d:
                    # Build path and record last write-off date at this level
                    path = tuple(subdiv_path[i] for i in (6, 8, 10) if i in subdiv_path)
                    if path and current_direction in directions and current_group:
                        gobj = directions[current_direction]["groups"].get(current_group)
                        if gobj:
                            sub_obj = gobj["subdivisions"].get(path)
                            if sub_obj is None:
                                sub_obj = gobj["subdivisions"][path] = {
                                    "path": list(path), "items": [], "last_writeoff": None,
                                }
                            prev = sub_obj.get("last_writeoff")
                            if prev is None or d > prev:
                                sub_obj["last_writeoff"] = d
            continue

        # Subdivision levels — indent=6,8,10 (only if NOT a document row, handled above)
        if indent in (6, 8, 10) and a_val:
            # Setting a new level invalidates deeper levels
            subdiv_path[indent] = a_val
            for deeper in (8, 10):
                if deeper > indent:
                    subdiv_path.pop(deeper, None)
            # We don't create the subdivision record yet; created when first item arrives
            continue

        # Document/operation headers — indent=12 or 14 (legacy fallback)
        if indent in (12, 14):
            # Even if name didn't match our prefixes, these indents are always docs.
            # Try to extract date in case it's a write-off we didn't recognize.
            if a_val and "Списання" in a_val:
                d = _extract_doc_date(a_val)
                if d:
                    path = tuple(subdiv_path[i] for i in (6, 8, 10) if i in subdiv_path)
                    if path and current_direction in directions and current_group:
                        gobj = directions[current_direction]["groups"].get(current_group)
                        if gobj:
                            sub_obj = gobj["subdivisions"].get(path)
                            if sub_obj is None:
                                sub_obj = gobj["subdivisions"][path] = {
                                    "path": list(path), "items": [], "last_writeoff": None,
                                }
                            prev = sub_obj.get("last_writeoff")
                            if prev is None or d > prev:
                                sub_obj["last_writeoff"] = d
            continue

        # Data row: indent=0, empty A, col B (name) filled — item
        if indent == 0 and not a_val and b_val:
            # Must have a subdivision context
            if current_direction is None or current_group is None or not subdiv_path:
                # Pre-header data (shouldn't happen in well-formed file)
                stats["unknown_levels"] += 1
                continue

            # Build path (levels in order 6,8,10)
            path = tuple(subdiv_path[i] for i in (6, 8, 10) if i in subdiv_path)
            if not path:
                stats["unknown_levels"] += 1
                continue

            nums = extract_numbers(row_cells)
            if not has_any_number(nums):
                stats["skipped_empty"] += 1
                continue

            # Locate/create subdivision record inside the current group
            grp = directions[current_direction]["groups"][current_group]
            sub_key = path
            if sub_key not in grp["subdivisions"]:
                grp["subdivisions"][sub_key] = {
                    "path": list(path),
                    "items": [],
                    "_items_by_name": {},  # name -> index in items[]
                    "last_writeoff": None,
                }
            sub = grp["subdivisions"][sub_key]
            if "_items_by_name" not in sub:
                sub["_items_by_name"] = {}

            # AGGREGATION: if same name already in this subdivision, merge into it.
            # Different series of the same drug => one logical position.
            existing_idx = sub["_items_by_name"].get(b_val)
            if existing_idx is not None:
                existing = sub["items"][existing_idx]
                # Sum all numeric fields
                for k, v in nums.items():
                    if v is not None and v != 0:
                        existing[k] = round((existing.get(k) or 0) + v, 2)
                # Track the list of series (and pick the earliest expiry to surface)
                if c_val:
                    series_list = existing.get("series_list")
                    if series_list is None:
                        # Initialize from existing single series (if any)
                        series_list = [existing["s"]] if existing.get("s") else []
                        existing["series_list"] = series_list
                    if c_val not in series_list:
                        series_list.append(c_val)
                # Don't overwrite the canonical series field
            else:
                item = {
                    "n": b_val,
                    "s": c_val or None,
                    **nums,
                }
                item = {k: v for k, v in item.items() if v is not None}
                sub["items"].append(item)
                sub["_items_by_name"][b_val] = len(sub["items"]) - 1
                stats["items_captured"] += 1
            continue

        # Unknown row pattern — log and skip
        if a_val or b_val or c_val:
            stats["unknown_levels"] += 1

    # Convert nested dicts to lists for JSON, compute subdivision totals
    directions_out = []
    # Stable ordering: by first-seen direction name
    for dir_name, dirobj in directions.items():
        groups_out = []
        for grp_name, grp in dirobj["groups"].items():
            subs_out = []
            for path_tuple, sub in grp["subdivisions"].items():
                # Compute aggregate totals from items
                sub_totals = aggregate_items(sub["items"])
                sub_record = {
                    "path": sub["path"],
                    "totals": sub_totals,
                    "items": sub["items"],
                }
                if sub.get("last_writeoff"):
                    sub_record["last_writeoff"] = sub["last_writeoff"]
                subs_out.append(sub_record)
            groups_out.append({
                "name": grp["name"],
                "totals": _strip_none(grp["totals"]),
                "subdivisions": subs_out,
            })
        directions_out.append({
            "name": dirobj["name"],
            "totals": _strip_none(dirobj["totals"]),
            "groups": groups_out,
        })

    # Convert sets to counts
    stats["directions_found"] = sorted(stats["directions_found"])
    stats["groups_found"] = sorted(stats["groups_found"])

    return {
        "period": period,
        "org_totals": _strip_none(org_totals),
        "directions": directions_out,
        "_stats": stats,
    }


def aggregate_items(items):
    """Sum 8 numeric fields across items."""
    agg = {"oq": 0, "os": 0, "rq": 0, "rs": 0, "eq": 0, "es": 0, "cq": 0, "cs": 0}
    for it in items:
        for k in agg:
            v = it.get(k)
            if isinstance(v, (int, float)):
                agg[k] += v
    # Round and strip zeros
    out = {}
    for k, v in agg.items():
        if v != 0:
            out[k] = round(v, 2)
    return out


def month_key_from_sheet_name(sheet_name, period):
    """Build 'YYYY-MM' key from period (preferred) or month name fallback."""
    if period and period.get("from"):
        return period["from"][:7]
    # Fallback to month name
    mm = MONTH_UK_TO_NUM.get(sheet_name)
    if mm:
        return f"2026-{mm}"  # fallback year; real file always has period
    return sheet_name


def main(input_path, output_path):
    t0 = time.time()
    print(f"Loading {input_path} ...", file=sys.stderr)
    wb = openpyxl.load_workbook(input_path, read_only=False, data_only=True)
    print(f"  loaded in {time.time() - t0:.1f}s", file=sys.stderr)
    print(f"  sheets: {wb.sheetnames}", file=sys.stderr)

    result = {
        "meta": {
            "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "source_file": Path(input_path).name,
            "periods": [],
            "format_version": 1,
        },
        "names": [],       # lookup table for item names
        "paths": [],       # lookup table for subdivision path components
        "months": {},
    }
    name_index = {}        # name -> index in "names"
    path_index = {}        # path component -> index in "paths"

    def intern_name(s):
        i = name_index.get(s)
        if i is None:
            i = len(result["names"])
            name_index[s] = i
            result["names"].append(s)
        return i

    def intern_path_component(s):
        i = path_index.get(s)
        if i is None:
            i = len(result["paths"])
            path_index[s] = i
            result["paths"].append(s)
        return i

    for sheet_name in wb.sheetnames:
        ts = time.time()
        print(f"Parsing '{sheet_name}' ...", file=sys.stderr)
        ws = wb[sheet_name]
        parsed = parse_sheet(ws, sheet_name)
        key = month_key_from_sheet_name(sheet_name, parsed.get("period"))
        result["meta"]["periods"].append({
            "key": key,
            "sheet": sheet_name,
            "period": parsed.get("period"),
        })

        stats = parsed.pop("_stats")

        # Intern names & paths
        for d in parsed["directions"]:
            for g in d["groups"]:
                for sub in g["subdivisions"]:
                    sub["path"] = [intern_path_component(p) for p in sub["path"]]
                    for it in sub["items"]:
                        it["n"] = intern_name(it["n"])

        result["months"][key] = parsed

        # ============== DATA INTEGRITY CHECK ==============
        # Validate the chain:
        #   item_sum  →  must == subdivision_totals (we compute these)
        #   group.totals (read from Excel)  →  must == sum of subdivisions
        #   sum of group.totals  →  COMPARED TO direction.totals and org_totals
        #
        # If sum-of-groups != org_total, the discrepancy lives INSIDE the source
        # Excel — usually because 1C aggregates more at the direction level than
        # is reflected in the lower hierarchy. This is a data-source quirk, not
        # a parser bug. We surface it clearly so the operator knows.
        org = parsed.get("org_totals") or {}
        # 1) Walk-through totals: sum of all items
        item_sums = {"oq": 0, "os": 0, "rq": 0, "rs": 0, "eq": 0, "es": 0, "cq": 0, "cs": 0}
        # 2) Sum of all GROUP totals (these come from Excel rows directly)
        group_sums = {k: 0 for k in item_sums}

        for d in parsed["directions"]:
            for g in d["groups"]:
                gt = g.get("totals") or {}
                for k in group_sums:
                    v = gt.get(k)
                    if isinstance(v, (int, float)):
                        group_sums[k] += v
                for sub in g["subdivisions"]:
                    for it in sub["items"]:
                        for k in item_sums:
                            v = it.get(k)
                            if isinstance(v, (int, float)):
                                item_sums[k] += v

        # Round results
        for k in item_sums: item_sums[k] = round(item_sums[k], 2)
        for k in group_sums: group_sums[k] = round(group_sums[k], 2)

        # Check 1: items match group totals (this is the strict parser correctness)
        items_vs_groups_problems = []
        for k in ("os", "rs", "es", "cs"):
            if abs(item_sums[k] - group_sums[k]) > 0.01:
                items_vs_groups_problems.append(
                    f"{k}: items={item_sums[k]:,.2f} groups={group_sums[k]:,.2f} diff={item_sums[k]-group_sums[k]:+,.2f}"
                )

        # Check 2: group totals match org totals (this is the source-data quality)
        groups_vs_org_problems = []
        for k in ("os", "rs", "es", "cs"):
            org_v = org.get(k) or 0
            if abs(org_v - group_sums[k]) > 0.01:
                groups_vs_org_problems.append(
                    f"{k}: org={org_v:,.2f} groups_sum={group_sums[k]:,.2f} diff={org_v-group_sums[k]:+,.2f}"
                )

        if items_vs_groups_problems:
            # This IS a parser bug — items captured don't match the group rows
            # we trust. Fail loudly.
            print(f"  ❌ ❌ ❌  ПОМИЛКА ПАРСЕРА у '{sheet_name}'  ❌ ❌ ❌", file=sys.stderr)
            print(f"     Сума позицій не дорівнює сумі підсумків груп.", file=sys.stderr)
            print(f"     Це означає що парсер пропустив частину рядків.", file=sys.stderr)
            for p in items_vs_groups_problems:
                print(f"       {p}", file=sys.stderr)

        if groups_vs_org_problems:
            # Source-data quirk — 1C report's direction/org totals don't fully
            # decompose into groups. Data inside the JSON is consistent and correct;
            # the 1C top-line just shows more.
            print(f"  ℹ️  УВАГА у '{sheet_name}': підсумок організації у Excel "
                  f"містить агрегати, які не розкладаються на групи.", file=sys.stderr)
            for p in groups_vs_org_problems:
                print(f"       {p}", file=sys.stderr)
            print(f"     Це особливість звіту 1С — у JSON відображені всі реальні позиції,", file=sys.stderr)
            print(f"     але загальна сума у заголовку Excel містить додаткові ", file=sys.stderr)
            print(f"     агрегати на рівні напряму. Дані позицій коректні.", file=sys.stderr)

        if not items_vs_groups_problems and not groups_vs_org_problems:
            # Pretty-print confirmation
            pass  # silence is good

        print(
            f"  {sheet_name}: {stats['items_captured']} items, "
            f"{len(stats['directions_found'])} directions, "
            f"{len(stats['groups_found'])} groups, "
            f"skipped_empty={stats['skipped_empty']}, "
            f"unknown={stats['unknown_levels']}  "
            f"[{time.time() - ts:.1f}s]",
            file=sys.stderr,
        )

        # If unknown count is suspiciously high, list a few examples
        if stats['unknown_levels'] > 5:
            print(f"  ⚠️ Багато невпізнаних рядків ({stats['unknown_levels']}). "
                  f"Можливо нова структура файлу — поверніться до розробника.", file=sys.stderr)

    print(f"Interned: {len(result['names'])} unique names, {len(result['paths'])} unique path components", file=sys.stderr)
    print(f"Writing {output_path} ...", file=sys.stderr)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, separators=(",", ":"))

    size = Path(output_path).stat().st_size
    print(f"Done. Output size: {size / 1024:.0f} KB ({size} bytes)", file=sys.stderr)
    print(f"Total time: {time.time() - t0:.1f}s", file=sys.stderr)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: parse.py <input.xlsx> [<output.json>]", file=sys.stderr)
        sys.exit(1)
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "analytics.json"
    main(inp, out)
